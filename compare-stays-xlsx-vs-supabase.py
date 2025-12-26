import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook


@dataclass(frozen=True)
class SupabaseConfig:
    url: str
    anon_key: str


def read_env_value(env_path: Path, key: str) -> Optional[str]:
    if not env_path.exists():
        return None
    for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line or line.lstrip().startswith("#"):
            continue
        if not line.startswith(key + "="):
            continue
        val = line.split("=", 1)[1].strip()
        if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
            val = val[1:-1]
        return val.strip()
    return None


def load_supabase_config(repo_root: Path) -> SupabaseConfig:
    env_path = repo_root / ".env.local"
    supabase_url = read_env_value(env_path, "VITE_SUPABASE_URL") or read_env_value(env_path, "SUPABASE_URL")
    anon_key = read_env_value(env_path, "VITE_SUPABASE_ANON_KEY") or read_env_value(env_path, "SUPABASE_ANON_KEY")

    if not supabase_url or not anon_key:
        raise RuntimeError(f"Missing VITE_SUPABASE_URL/VITE_SUPABASE_ANON_KEY in {env_path}")

    return SupabaseConfig(url=supabase_url.rstrip("/"), anon_key=anon_key)


def normalize_id(value: object) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return None

    # Excel may coerce IDs to scientific notation if they look numeric; keep only if it matches expected hex-ish id.
    # Stays IDs we saw look like 24-hex Mongo-ish.
    m = re.search(r"\b[0-9a-fA-F]{24}\b", s)
    if m:
        return m.group(0).lower()

    # Some exports may include prefixes or other formats; return raw if it has at least 8 chars.
    if len(s) >= 8:
        return s
    return None


def normalize_reserve_code(value: object) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return None
    s = s.upper()
    # Reserva codes in Stays exports are short alphanumeric codes (e.g. EU26J, DQ52J)
    if re.fullmatch(r"[A-Z0-9]{3,12}", s):
        return s
    return None


def extract_xlsx_reserva_codes(xlsx_path: Path) -> Set[str]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    out: Set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        headers = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(header)]
        if "Reserva" not in headers:
            continue
        col_idx = headers.index("Reserva")
        for row in rows:
            if col_idx >= len(row):
                continue
            code = normalize_reserve_code(row[col_idx])
            if code:
                out.add(code)
        # Typically only one sheet matters; stop once found.
        if out:
            break

    return out


def extract_ids_by_column(xlsx_path: Path) -> Dict[Tuple[str, str], Set[str]]:
    """Return {(sheetName, columnHeader): set(ids)} for every column.

    We intentionally extract per-column so we can auto-detect which column represents reservation IDs.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    out: Dict[Tuple[str, str], Set[str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        headers = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(header)]
        col_sets: List[Set[str]] = [set() for _ in headers]

        for row in rows:
            for i, cell in enumerate(row[: len(headers)]):
                nid = normalize_id(cell)
                if nid and re.fullmatch(r"[0-9a-f]{24}", nid):
                    col_sets[i].add(nid)

        for i, ids in enumerate(col_sets):
            if ids:
                out[(sheet_name, headers[i])] = ids

    return out


def pick_best_column(xlsx_path: Path, supabase_ids: Set[str]) -> Tuple[Set[str], str]:
    cols = extract_ids_by_column(xlsx_path)
    best_ids: Set[str] = set()
    best_key: Optional[Tuple[str, str]] = None
    best_overlap = -1

    for key, ids in cols.items():
        overlap = len(ids & supabase_ids)
        # Prefer higher overlap; tie-breaker: closer to 498-ish.
        if overlap > best_overlap or (overlap == best_overlap and abs(len(ids) - 498) < abs(len(best_ids) - 498)):
            best_overlap = overlap
            best_ids = ids
            best_key = key

    if not best_key:
        return set(), "<no columns with 24-hex ids found>"

    sheet, col = best_key
    return best_ids, f"{sheet}::{col} (overlapWithSupabase={best_overlap}, idsInColumn={len(best_ids)})"


def fetch_supabase_external_ids(
    cfg: SupabaseConfig,
    check_in_from: str,
    check_in_to: str,
    only_imported: bool = True,
) -> List[str]:
    headers = {
        "apikey": cfg.anon_key,
        "Authorization": f"Bearer {cfg.anon_key}",
    }

    filters = [
        f"check_in=gte.{check_in_from}",
        f"check_in=lte.{check_in_to}",
    ]
    if only_imported:
        filters.append("external_id=not.is.null")

    select = "external_id"
    page_size = 1000
    offset = 0
    out: List[str] = []

    while True:
        qs_parts = [
            f"select={select}",
            f"order=external_id.asc",
            f"limit={page_size}",
            f"offset={offset}",
            *filters,
        ]
        uri = f"{cfg.url}/rest/v1/reservations?" + "&".join(qs_parts)
        resp = requests.get(uri, headers=headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, list):
            raise RuntimeError(f"Unexpected response type: {type(data)}")

        got = 0
        for row in data:
            ext = normalize_id(row.get("external_id"))
            if ext and re.fullmatch(r"[0-9a-f]{24}", ext.lower()):
                out.append(ext.lower())
                got += 1

        if len(data) < page_size:
            break
        offset += page_size

    # De-dup (should already be unique, but keep stable)
    return sorted(set(out))


def extract_reserve_code_from_url(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        reserve = qs.get("reserve", [None])[0]
        return normalize_reserve_code(reserve)
    except Exception:
        return None


def fetch_supabase_reserve_codes(
    cfg: SupabaseConfig,
    check_in_from: str,
    check_in_to: str,
    only_imported: bool = True,
) -> Tuple[Dict[str, str], Set[str]]:
    """Return (reserveCode -> reservationId, set(reserveCodes))."""
    headers = {
        "apikey": cfg.anon_key,
        "Authorization": f"Bearer {cfg.anon_key}",
    }

    filters = [
        f"check_in=gte.{check_in_from}",
        f"check_in=lte.{check_in_to}",
    ]
    if only_imported:
        filters.append("external_id=not.is.null")

    select = "id,external_url"
    page_size = 1000
    offset = 0
    mapping: Dict[str, str] = {}
    codes: Set[str] = set()

    while True:
        qs_parts = [
            f"select={select}",
            f"order=check_in.asc",
            f"limit={page_size}",
            f"offset={offset}",
            *filters,
        ]
        uri = f"{cfg.url}/rest/v1/reservations?" + "&".join(qs_parts)
        resp = requests.get(uri, headers=headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, list):
            raise RuntimeError(f"Unexpected response type: {type(data)}")

        for row in data:
            rid = row.get("id")
            code = extract_reserve_code_from_url(row.get("external_url"))
            if rid and code:
                codes.add(code)
                # If duplicates happen, keep the first and still collect codes.
                mapping.setdefault(code, str(rid))

        if len(data) < page_size:
            break
        offset += page_size

    return mapping, codes


def main() -> int:
    workspace_root = Path(__file__).resolve().parents[1]
    repo_root = Path(__file__).resolve().parent

    xlsx = workspace_root / "planilhas auditoria" / "2025-12-01_2025-12-31_propX_ownerX.xlsx"
    if not xlsx.exists():
        print(f"ERROR: XLSX not found: {xlsx}")
        return 2

    cfg = load_supabase_config(repo_root)

    xlsx_reserva = extract_xlsx_reserva_codes(xlsx)
    mapping, supa_reserva = fetch_supabase_reserve_codes(cfg, "2025-12-01", "2025-12-31", only_imported=True)

    only_in_supabase = sorted(supa_reserva - xlsx_reserva)
    only_in_xlsx = sorted(xlsx_reserva - supa_reserva)

    print(f"XLSX_RESERVA_CODES={len(xlsx_reserva)}")
    print(f"SUPABASE_RESERVA_CODES={len(supa_reserva)}")
    print(f"ONLY_IN_SUPABASE={len(only_in_supabase)}")
    print(f"ONLY_IN_XLSX={len(only_in_xlsx)}")

    if only_in_supabase:
        print("\nEXTRA_IN_SUPABASE (Reserva codes):")
        for code in only_in_supabase[:50]:
            rid = mapping.get(code)
            suffix = f" -> reservationId={rid}" if rid else ""
            print(f"{code}{suffix}")

    if only_in_xlsx:
        print("\nMISSING_IN_SUPABASE (Reserva codes):")
        for code in only_in_xlsx[:50]:
            print(code)

    return 0 if (not only_in_supabase and not only_in_xlsx) else 1


if __name__ == "__main__":
    raise SystemExit(main())
